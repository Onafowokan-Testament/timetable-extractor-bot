import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import matplotlib.pyplot as plt
import telebot


bot = telebot.TeleBot("bot_key")


def search_courses(df, courses):
    found_courses = []

    for index, row in df.iterrows():
        building = row['BUILDING']
        room = row.iloc[1]

        for col in df.columns[2:]:

            if '_' in col:
                time_slot = col.split('_')
                start_time, end_time = time_slot[0], time_slot[1]
            else:
                start_time, end_time = None, None

            subjects = str(row[col]).split('/') if pd.notna(row[col]) else []
            for subject in subjects:
                stripped_subject = subject.strip()
                if stripped_subject in courses:
                    found_courses.append({
                        'Course': stripped_subject,
                        'Building': building,
                        'Room': room,
                        'Time': f"{start_time}-{end_time}" if start_time and end_time else None,
                        'Day': df['Day'][index]
                    })

    found_courses_df = pd.DataFrame(found_courses)
    return found_courses_df


def generate_timetable_pivot(found_courses_all_days):
    days_present = found_courses_all_days['Day'].unique()

    desired_time_slots = [
        '8am-9am', '9am-10am', '10am-11am', '11am-12noon',
        '12noon-1pm', '1pm-2pm', '2pm-3pm', '3pm-4pm',
        '4pm-5pm', '5pm-6pm', '6pm-7pm'
    ]

    timetable_pivot = found_courses_all_days.pivot_table(
        index='Day',
        columns='Time',
        values='Course',
        aggfunc=lambda x: ', '.join(
            x.dropna().astype(str)) if x.notna().any() else ''
    )

    timetable_pivot = timetable_pivot.reindex(
        index=days_present, fill_value='')
    timetable_pivot = timetable_pivot.reindex(
        columns=desired_time_slots, fill_value='')

    return timetable_pivot


@bot.message_handler(commands=['start'])
def handle_start(message):
    bot.reply_to(
        message, "Welcome to the Timetable Bot! Please use the /courses command to input your courses.")


@bot.message_handler(commands=['courses'])
def handle_courses(message):
    user_id = message.from_user.id
    courses_list = []

    bot.reply_to(
        message, "Please enter the number of courses you want to add:")
    bot.register_next_step_handler(
        message, get_courses_count, user_id, courses_list)


def get_courses_count(message, user_id, courses_list):
    try:
        num_courses = int(message.text)
        bot.reply_to(
            message, f"Great! Please enter your {num_courses} courses one by one. i.e courses shoul be in capital letters and no space e.g ECN311")
        bot.register_next_step_handler(
            message, get_courses, user_id, num_courses, courses_list)
    except ValueError:
        bot.reply_to(message, "Please enter a valid number.")


def get_courses(message, user_id, num_courses, courses_list):
    courses_list.append(message.text.upper())

    if len(courses_list) < num_courses:
        bot.reply_to(
            message, f"Course {len(courses_list)} recorded. Enter the next course:")
        bot.register_next_step_handler(
            message, get_courses, user_id, num_courses, courses_list)
    else:
        bot.reply_to(
            message, "Courses recorded successfully! Generating your timetable...")
        generate_and_send_timetable(user_id, courses_list)


def generate_and_send_timetable(user_id, courses_list):

    monday_timetable = pd.read_csv('timetable/monday.csv')
    tuesday_timetable = pd.read_csv('timetable/tuesday.csv')
    wednesday_timetable = pd.read_csv('timetable/wednesday.csv')
    thursday_timetable = pd.read_csv('timetable/thursday.csv')
    friday_timetable = pd.read_csv('timetable/friday.csv')

    # Rename columns to make them universal
    columns_mapping = {
        'BUILDING\r\nB': 'BUILDING',
        'Unnamed: 14': 'Day'
    }
    timetables = [monday_timetable, tuesday_timetable,
                  wednesday_timetable, thursday_timetable, friday_timetable]

    for df in timetables:
        df.rename(columns=columns_mapping, inplace=True)

    # Add 'Day' column to each dataframe
    days_of_week = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    for df, day in zip(timetables, days_of_week):
        df['Day'] = day

    # Search for courses in each timetable
    found_courses_all_days = pd.DataFrame()
    for df in timetables:
        found_courses_df = search_courses(df, courses_list)
        found_courses_all_days = pd.concat(
            [found_courses_all_days, found_courses_df], ignore_index=True)

    timetable_pivot = generate_timetable_pivot(found_courses_all_days)

    wb = Workbook()
    ws = wb.active

    font_size = 12
    bold_font = Font(bold=True, size=font_size)

    ws.append([''] + list(timetable_pivot.columns))
    for cell in ws[1]:
        cell.font = bold_font

    for day, row in timetable_pivot.iterrows():
        ws.append([day] + list(row))

        ws.row_dimensions[ws.max_row].height = 40
        for cell in ws.iter_cols(min_col=2, max_col=ws.max_column, min_row=ws.max_row, max_row=ws.max_row):
            cell[0].alignment = Alignment(
                wrap_text=True, vertical='center', horizontal='center')

    wb.save("timetable.xlsx")

    found_courses_all_days.to_csv('found_courses.csv', index=False)

    bot.send_document(user_id, open('timetable.xlsx', 'rb'),
                      caption='Your Timetable')


bot.polling()
